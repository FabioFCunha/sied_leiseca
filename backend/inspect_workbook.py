from pathlib import Path

from openpyxl import load_workbook


def first_non_empty_row(sheet, max_rows=20):
    for row in sheet.iter_rows(min_row=1, max_row=max_rows, values_only=True):
        values = [str(value).strip() if value is not None else "" for value in row]
        if any(values):
            return values
    return []


def main():
    path = Path(r"C:\Users\fferreira\Downloads\AGENTES.xlsx")
    workbook = load_workbook(path, read_only=True, data_only=True)
    print(f"Arquivo: {path}")
    for sheet in workbook.worksheets:
        header = first_non_empty_row(sheet)
        non_empty_rows = 0
        samples = []
        for row in sheet.iter_rows(values_only=True):
            values = [value for value in row if value not in (None, "")]
            if values:
                non_empty_rows += 1
                if len(samples) < 3:
                    samples.append([str(value) if value is not None else "" for value in row[:12]])
        print("\n---")
        print(f"Aba: {sheet.title}")
        print(f"Dimensao: {sheet.max_row} linhas x {sheet.max_column} colunas")
        print(f"Linhas com dados: {non_empty_rows}")
        print(f"Cabecalho provável: {header[:60]}")
        print("Amostras:")
        for sample in samples:
            print(sample)


if __name__ == "__main__":
    main()
