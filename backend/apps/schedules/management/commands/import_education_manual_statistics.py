from datetime import date
from pathlib import Path
import re
import unicodedata

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from apps.accounts.models import User
from apps.schedules.models import EducationAction, EducationReport


FIELD_MAP = {
    "1 ABORDADOS": "approach",
    "1 1 ABORDADOS PALESTRAS": "approached_lectures",
    "1 2 ABORDADOS ACOES": "approached_actions",
    "2 PALESTRAS REALIZADAS INICIO 4 2011": "lectures",
    "2 1 ESCOLAS": "schools",
    "2 2 UNIVERSIDADES": "universities",
    "2 3 EMPRESAS": "companies",
    "3 ACOES": "educational_actions",
    "3 1 BARES": "bars",
    "3 2 PEDAGIO": "tolls",
    "3 3 ESPORTES": "sports",
    "3 4 PRAIA": "beach",
    "3 5 EVENTOS": "events",
    "3 6 SHOPPING": "shopping",
    "3 7 ACAO SOCIAL": "social_actions",
    "3 8 OUTROS": "other_actions",
    "4 MATERIAIS DE DIVULGACAO": "publicity_materials",
    "2 4 CERTIFICADOS ENTREGUES": "distributed_certificates",
    "3 REVISTINHA SOPRINHO": "gibis",
}


def clean(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def key(value):
    text = unicodedata.normalize("NFKD", clean(value)).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^A-Z0-9]+", " ", text.upper()).strip()


def number(value):
    if value in (None, ""):
        return 0
    return int(float(value))


class Command(BaseCommand):
    help = "Importa o balanco manual anual da educacao para as estatisticas dos relatorios tecnicos."

    def add_arguments(self, parser):
        parser.add_argument("workbook")
        parser.add_argument("--sheet", default="Plan1")
        parser.add_argument("--user-email", default="")
        parser.add_argument("--dry-run", action="store_true")

    @transaction.atomic
    def handle(self, *args, **options):
        path = Path(options["workbook"])
        if not path.exists():
            raise CommandError(f"Arquivo nao encontrado: {path}")

        user = self.get_user(options["user_email"])
        workbook = load_workbook(path, data_only=True)
        if options["sheet"] not in workbook.sheetnames:
            raise CommandError(f"Aba {options['sheet']} nao encontrada. Abas: {', '.join(workbook.sheetnames)}")

        sheet = workbook[options["sheet"]]
        years = self.year_columns(sheet)
        rows = self.metric_rows(sheet)
        created = 0
        updated = 0

        for year, column in years:
            report, report_created = EducationReport.objects.update_or_create(
                source_id=f"manual-statistics:{year}",
                defaults={
                    "source": EducationReport.Source.IMPORTED,
                    "operation_date": date(year, 4, 30) if year == 2026 else date(year, 12, 31),
                    "team": "Historico",
                    "management_name": "Balanco manual da Educacao",
                    "status": EducationReport.ReportStatus.SUBMITTED,
                    "created_by": user,
                    "occurrence_observation": f"Numeros importados da planilha {path.name}.",
                },
            )
            payload = {
                "agenda": None,
                "place_action": "Balanco anual da Educacao",
                "type_action": "Historico importado",
                "type_audience": "Consolidado anual",
                "institution_name": "Educacao",
            }
            for metric_key, row in rows.items():
                field = FIELD_MAP.get(metric_key)
                if field:
                    payload[field] = number(sheet.cell(row=row, column=column).value)

            EducationAction.objects.update_or_create(
                source_id=f"manual-statistics:{year}:aggregate",
                defaults={"report": report, **payload},
            )
            created += int(report_created)
            updated += int(not report_created)

        if options["dry_run"]:
            transaction.set_rollback(True)
            self.stdout.write(self.style.WARNING("Dry-run: nada foi gravado no banco."))

        self.stdout.write(
            self.style.SUCCESS(
                f"Importacao concluida: {created} anos criados, {updated} anos atualizados, "
                f"{len(years)} acoes consolidadas gravadas."
            )
        )

    def get_user(self, email):
        queryset = User.objects.all()
        if email:
            user = queryset.filter(email__iexact=email).first()
            if not user:
                raise CommandError(f"Usuario nao encontrado: {email}")
            return user
        user = queryset.filter(role=User.Role.ADMIN).order_by("id").first() or queryset.filter(is_superuser=True).order_by("id").first()
        if not user:
            raise CommandError("Nenhum administrador encontrado. Informe --user-email.")
        return user

    def year_columns(self, sheet):
        columns = []
        for column in range(1, sheet.max_column + 1):
            value = sheet.cell(row=2, column=column).value
            if isinstance(value, int) and 2000 <= value <= 2100:
                columns.append((value, column))
        if not columns:
            raise CommandError("Nao encontrei anos na linha 2 da planilha.")
        return columns

    def metric_rows(self, sheet):
        rows = {}
        for row in range(1, sheet.max_row + 1):
            label = key(sheet.cell(row=row, column=2).value)
            if label:
                rows[label] = row
        missing = sorted(set(FIELD_MAP) - set(rows))
        if missing:
            raise CommandError(f"Indicadores nao encontrados na planilha: {', '.join(missing)}")
        return rows
