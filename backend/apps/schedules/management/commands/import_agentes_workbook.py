from datetime import datetime, time, timedelta
from pathlib import Path
import re

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from apps.accounts.models import User
from apps.schedules.models import (
    ActionType,
    Agent,
    Agenda,
    AgendaHistory,
    AgendaMaterial,
    Chief,
    Kit,
    Material,
    Municipality,
    Neighborhood,
    Sector,
    Support,
    Team,
    Vehicle,
)
from apps.schedules.views import snapshot_for


STATUS_MAP = {
    "AGUARDANDO": Agenda.Status.PENDING,
    "CONFIRMADO": Agenda.Status.APPROVED,
    "NAO CONFIRMADO": Agenda.Status.CANCELLED,
    "NÃO CONFIRMADO": Agenda.Status.CANCELLED,
    "CANCELADO": Agenda.Status.CANCELLED,
    "CANCELADA": Agenda.Status.CANCELLED,
    "CONCLUIDO": Agenda.Status.COMPLETED,
    "CONCLUÍDO": Agenda.Status.COMPLETED,
    "CONCLUIDA": Agenda.Status.COMPLETED,
    "CONCLUÍDA": Agenda.Status.COMPLETED,
}


def clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalized(value):
    return re.sub(r"\s+", " ", clean(value)).strip()


def get_by_name(model, name, source_id=""):
    name = normalized(name)
    if model in (Kit, Support):
        name = name.upper()
    if not name:
        return None
    found = model.objects.filter(name__iexact=name).first()
    if found:
        if source_id and not found.source_id:
            found.source_id = source_id
            found.save(update_fields=["source_id"])
        return found
    return model.objects.create(name=name, source_id=source_id)


def import_lookup_sheet(workbook, sheet_name, model, name_header, extra=None):
    if sheet_name not in workbook.sheetnames:
        return 0
    sheet = workbook[sheet_name]
    headers = [normalized(cell.value).upper() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    name_index = headers.index(name_header.upper())
    id_index = headers.index("ID") if "ID" in headers else 0
    count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = normalized(row[name_index] if name_index < len(row) else "")
        if model in (Kit, Support):
            name = name.upper()
        if not name:
            continue
        item = get_by_name(model, name, clean(row[id_index] if id_index < len(row) else ""))
        if extra:
            extra(item, row, headers)
        count += 1
    return count


def parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    text = clean(value)
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def parse_time(value):
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    text = clean(value)
    if not text:
        return time(9, 0)
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            pass
    return time(9, 0)


def parse_int(value):
    text = clean(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def row_dict(headers, row):
    return {
        headers[index]: row[index] if index < len(row) else None
        for index in range(len(headers))
    }


def split_agents(value):
    text = normalized(value)
    if not text:
        return []
    parts = re.split(r"\s+-\s+|;|,", text)
    return [part.strip() for part in parts if part.strip() and part.strip() != "0"]


class Command(BaseCommand):
    help = "Importa cadastros auxiliares e agendas do arquivo AGENTES.xlsx."

    def add_arguments(self, parser):
        parser.add_argument("path", nargs="?", default=r"C:\Users\fferreira\Downloads\AGENTES.xlsx")

    @transaction.atomic
    def handle(self, *args, **options):
        path = Path(options["path"])
        if not path.exists():
            raise CommandError(f"Arquivo não encontrado: {path}")

        workbook = load_workbook(path, read_only=True, data_only=True)
        admin = User.objects.filter(role=User.Role.ADMIN).first()
        if not admin:
            raise CommandError("Crie/importe um administrador antes de importar agendas.")

        counts = {
            "kits": import_lookup_sheet(workbook, "KIT", Kit, "KIT"),
            "vehicles": import_lookup_sheet(workbook, "VIATURA", Vehicle, "VIATURA"),
            "supports": import_lookup_sheet(workbook, "APOIO", Support, "APOIO"),
            "teams": import_lookup_sheet(workbook, "EQUIPE", Team, "EQUIPE"),
            "agents": import_lookup_sheet(workbook, "AGENTES", Agent, "AGENTES"),
            "municipalities": import_lookup_sheet(workbook, "MUNICIPIO", Municipality, "MUNICÍPIO"),
            "neighborhoods": import_lookup_sheet(workbook, "BAIRRO", Neighborhood, "BAIRRO"),
            "action_types": import_lookup_sheet(workbook, "TIPO DE ACAO", ActionType, "TIPO DE AÇÃO"),
        }

        def save_chief(item, row, headers):
            phone_index = headers.index("TELEFONE") if "TELEFONE" in headers else None
            if phone_index is not None:
                item.phone = normalized(row[phone_index] if phone_index < len(row) else "")
                item.save(update_fields=["phone"])

        counts["chiefs"] = import_lookup_sheet(workbook, "CHEFE", Chief, "CHEFES DE EQUIPE", save_chief)
        counts["agendas"] = self.import_agendas(workbook, admin)

        self.stdout.write(self.style.SUCCESS(f"Importação concluída: {counts}"))

    def import_agendas(self, workbook, admin):
        if "DADOS" not in workbook.sheetnames:
            raise CommandError("Aba DADOS não encontrada.")
        sheet = workbook["DADOS"]
        headers = [normalized(cell.value).upper() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data = row_dict(headers, row)
            source_id = normalized(data.get("ID"))
            agenda_date = parse_date(data.get("DATA"))
            if not source_id or not agenda_date:
                continue

            start_time = parse_time(data.get("HORÁRIO"))
            end_dt = datetime.combine(agenda_date, start_time) + timedelta(hours=1)
            vehicle = get_by_name(Vehicle, data.get("VIATURA"))
            team = get_by_name(Team, data.get("EQUIPE"))
            chief = get_by_name(Chief, data.get("CHEFE"))
            support_1 = get_by_name(Support, data.get("APOIO 1"))
            support_2 = get_by_name(Support, data.get("APOIO 2"))
            action_type = get_by_name(ActionType, data.get("TIPO DE AÇÃO"))
            municipality = get_by_name(Municipality, data.get("MUNICÍPIO"))
            neighborhood = get_by_name(Neighborhood, data.get("BAIRRO"))
            sector, _ = Sector.objects.get_or_create(
                name=normalized(data.get("EQUIPE")) or "Sem equipe",
                defaults={"description": "Criado pela importação do Excel"},
            )

            title = normalized(data.get("TIPO DE AÇÃO")) or normalized(data.get("INSTITUIÇÃO/LOCAL")) or source_id
            agenda, created = Agenda.objects.update_or_create(
                source_id=source_id,
                defaults={
                    "title": title[:180],
                    "description": normalized(data.get("OBSERVAÇÃO")) or f"Agenda importada {source_id}",
                    "date": agenda_date,
                    "start_time": start_time,
                    "end_time": end_dt.time().replace(second=0, microsecond=0),
                    "location": normalized(data.get("LOCAL")) or normalized(data.get("INSTITUIÇÃO/LOCAL")) or "-",
                    "vehicle": normalized(data.get("VIATURA")),
                    "vehicle_ref": vehicle,
                    "team_name": normalized(data.get("EQUIPE")),
                    "team_ref": team,
                    "chief_name": normalized(data.get("CHEFE")),
                    "chief_ref": chief,
                    "team_phone": normalized(data.get("TELEFONE")),
                    "agents": normalized(data.get("AGENTES")),
                    "support_1": normalized(data.get("APOIO 1")),
                    "support_1_ref": support_1,
                    "support_2": normalized(data.get("APOIO 2")),
                    "support_2_ref": support_2,
                    "action_type": normalized(data.get("TIPO DE AÇÃO")),
                    "action_type_ref": action_type,
                    "institution_location": normalized(data.get("INSTITUIÇÃO/LOCAL")),
                    "quantity": parse_int(data.get("QTD")),
                    "schedule_text": normalized(data.get("HORÁRIO")),
                    "address": normalized(data.get("ENDEREÇO")),
                    "neighborhood": normalized(data.get("BAIRRO")),
                    "neighborhood_ref": neighborhood,
                    "city": normalized(data.get("MUNICÍPIO")),
                    "municipality_ref": municipality,
                    "external_responsible": normalized(data.get("RESPONSÁVEL")),
                    "external_responsible_phone": normalized(data.get("TELEFONE DO RESPONSÁVEL")),
                    "external_email": normalized(data.get("E-MAIL")),
                    "audience": normalized(data.get("PÚBLICO")),
                    "activity_type": normalized(data.get("TIPO")),
                    "status": STATUS_MAP.get(normalized(data.get("STATUS")).upper(), Agenda.Status.PENDING),
                    "notes": normalized(data.get("OBSERVAÇÃO")),
                    "responsible": admin,
                    "sector": sector,
                    "created_by": admin,
                },
            )

            agenda.agents_ref.set(
                [get_by_name(Agent, agent_name) for agent_name in split_agents(data.get("AGENTES"))]
            )
            self.import_materials(agenda, data)
            if created:
                AgendaHistory.objects.create(
                    agenda=agenda,
                    changed_by=admin,
                    action="IMPORTACAO_EXCEL",
                    snapshot=snapshot_for(agenda),
                )
            count += 1
        return count

    def import_materials(self, agenda, data):
        AgendaMaterial.objects.filter(agenda=agenda).delete()
        for position in range(1, 8):
            kit_name = normalized(data.get(f"KIT {position}"))
            quantity = parse_int(data.get(f"QTD {position}"))
            material_name = normalized(data.get(f"MATERIAL{position}"))
            if not kit_name and not material_name and quantity is None:
                continue
            AgendaMaterial.objects.create(
                agenda=agenda,
                position=position,
                kit=get_by_name(Kit, kit_name),
                material=get_by_name(Material, material_name),
                quantity=quantity,
            )
