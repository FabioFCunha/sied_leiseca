from pathlib import Path
import re
import unicodedata

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from apps.schedules.models import Agent, Chief, Support, Team


TEAM_NAMES = {
    "ALFA": "ALFA",
    "BRAVO": "BRAVO",
    "CHARLIE": "CHARLIE",
    "DELTA": "DELTA",
    "ECHO": "ECHO",
    "FOX": "FOX",
    "GOLF": "GOLF",
    "HOTEL": "HOTEL",
}


def clean(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def key(value):
    text = unicodedata.normalize("NFKD", clean(value)).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^A-Z0-9]+", " ", text.upper()).strip()


def team_name(value):
    normalized = key(value)
    return TEAM_NAMES.get(normalized, clean(value).upper())


def source_id(prefix, team, name):
    return f"education-staff:{prefix}:{key(team)}:{key(name)}"[:80]


def get_by_name(model, name, defaults=None):
    defaults = defaults or {}
    found = model.objects.filter(name__iexact=name).first()
    if found:
        changed = []
        for field, value in defaults.items():
            if value is not None and getattr(found, field) != value:
                setattr(found, field, value)
                changed.append(field)
        if changed:
            found.save(update_fields=changed)
        return found, False
    return model.objects.create(name=name, **defaults), True


class Command(BaseCommand):
    help = "Importa equipes, chefes e agentes do arquivo Endereco do Efetivo da Educacao.xlsx."

    def add_arguments(self, parser):
        parser.add_argument("workbook")
        parser.add_argument("--sheet", default="Plan1")
        parser.add_argument("--dry-run", action="store_true")
        parser.add_argument(
            "--deactivate-missing-agents",
            action="store_true",
            help="Marca como inativos os agentes que nao aparecem na planilha importada.",
        )
        parser.add_argument(
            "--replace-supports",
            action="store_true",
            help="Apaga os apoios existentes e recria os apoios da planilha do efetivo.",
        )

    @transaction.atomic
    def handle(self, *args, **options):
        path = Path(options["workbook"])
        if not path.exists():
            raise CommandError(f"Arquivo nao encontrado: {path}")

        workbook = load_workbook(path, data_only=True)
        if options["sheet"] not in workbook.sheetnames:
            raise CommandError(f"Aba {options['sheet']} nao encontrada. Abas: {', '.join(workbook.sheetnames)}")

        rows = self.parse_rows(workbook[options["sheet"]])

        counts = {
            "teams_created": 0,
            "teams_updated": 0,
            "chiefs_created": 0,
            "chiefs_updated": 0,
            "agents_created": 0,
            "agents_updated": 0,
            "agents_deactivated": 0,
            "supports_created": 0,
            "supports_updated": 0,
            "supports_deleted": 0,
            "skipped": 0,
        }
        official_agent_names = set()

        if options["replace_supports"]:
            counts["supports_deleted"], _ = Support.objects.all().delete()

        for row in rows:
            team, created = get_by_name(Team, row["team"], {"source_id": source_id("team", row["team"], row["team"])})
            counts["teams_created" if created else "teams_updated"] += 1

            if "CHEFE" in key(row["role"]):
                _, created = get_by_name(
                    Chief,
                    row["name"],
                    {
                        "source_id": source_id("chief", row["team"], row["name"]),
                        "team": team,
                        "role": row["role"],
                        "address": row["address"],
                    },
                )
                counts["chiefs_created" if created else "chiefs_updated"] += 1
            elif row["name"]:
                if "APOIO" in key(row["role"]):
                    _, support_created = get_by_name(
                        Support,
                        row["name"],
                        {
                            "source_id": source_id("support", row["team"], row["name"]),
                            "team": team,
                            "role": row["role"],
                            "address": row["address"],
                            "is_active": True,
                        },
                    )
                    counts["supports_created" if support_created else "supports_updated"] += 1
                else:
                    official_agent_names.add(row["name"])
                    _, created = get_by_name(
                        Agent,
                        row["name"],
                        {
                            "source_id": source_id("agent", row["team"], row["name"]),
                            "team": team,
                            "role": row["role"],
                            "address": row["address"],
                            "is_active": True,
                        },
                    )
                    counts["agents_created" if created else "agents_updated"] += 1
            else:
                counts["skipped"] += 1

        if options["deactivate_missing_agents"]:
            counts["agents_deactivated"] = Agent.objects.exclude(name__in=official_agent_names).filter(is_active=True).update(is_active=False)

        if options["dry_run"]:
            transaction.set_rollback(True)
            self.stdout.write(self.style.WARNING("Dry-run: nada foi gravado no banco."))
        self.stdout.write(
            self.style.SUCCESS(
                "Importacao concluida: "
                f"{counts['teams_created']} equipes criadas, {counts['teams_updated']} atualizadas; "
                f"{counts['chiefs_created']} chefes criados, {counts['chiefs_updated']} atualizados; "
                f"{counts['agents_created']} agentes criados, {counts['agents_updated']} atualizados; "
                f"{counts['agents_deactivated']} agentes inativados; "
                f"{counts['supports_deleted']} apoios apagados, {counts['supports_created']} criados, "
                f"{counts['supports_updated']} atualizados; "
                f"{counts['skipped']} ignorados."
            )
        )

    def parse_rows(self, sheet):
        rows = []
        current_team = ""
        for values in sheet.iter_rows(values_only=True):
            name = clean(values[0] if len(values) > 0 else "")
            role = clean(values[1] if len(values) > 1 else "")
            address = clean(values[2] if len(values) > 2 else "")
            if name and not role and not address:
                current_team = team_name(name)
                continue
            if not name or not current_team:
                continue
            rows.append(
                {
                    "team": current_team,
                    "name": name.upper(),
                    "role": role.upper(),
                    "address": address.upper(),
                }
            )
        return rows
