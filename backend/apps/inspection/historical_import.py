import hashlib
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from apps.inspection.models import (
    HISTORICAL_CUTOFF_DATE,
    InspectionHistoricalImportBatch,
    InspectionHistoricalStatistic,
    HistoricalSourceType,
    HistoricalTaxonomyEra,
)


MONTH_NAMES = {
    "JANEIRO": 1,
    "FEVEREIRO": 2,
    "MARCO": 3,
    "MARÇO": 3,
    "ABRIL": 4,
    "MAIO": 5,
    "JUNHO": 6,
    "JULHO": 7,
    "AGOSTO": 8,
    "SETEMBRO": 9,
    "OUTUBRO": 10,
    "NOVEMBRO": 11,
    "DEZEMBRO": 12,
}

TEAM_PATTERN = re.compile(r"^\s*EQUIPE\s+(.+?)\s*$", re.IGNORECASE)
DAILY_SHEET_PATTERN = re.compile(r"^D(?P<day>\d{1,2})$")
EXPLICIT_MONTH_YEAR_PATTERN = re.compile(
    r"(?:M[EÊ]S\s+DE|ACUMULADO\s+DO\s+M[EÊ]S)\s+(?P<month>[A-ZÃÇ]+)\s+(?P<year>20\d{2})"
)


def _normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _normalize_sheet_name(value):
    return _normalize_text(value).upper()


def _int_or_none(value):
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def _decimal_or_none(value):
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None


def compute_sha256(file_path: Path) -> str:
    digest = hashlib.sha256()
    with file_path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


@dataclass
class ParsedHistoricalRow:
    source_type: str
    taxonomy_era: str
    source_sheet: str
    source_row: int
    reference_date: date | None
    reference_year: int | None
    reference_month: int | None
    team: str
    source_team_label: str
    source_workbook_label: str
    metrics: dict


class InspectionHistoricalWorkbookParser:
    def __init__(self, file_path):
        self.file_path = Path(file_path)
        self.workbook_label = self.file_path.name
        self.workbook_values = None
        self.workbook_formulas = None

    def parse(self):
        if not self.file_path.exists():
            raise FileNotFoundError(f"Arquivo nao encontrado: {self.file_path}")

        file_sha256 = compute_sha256(self.file_path)
        file_size = self.file_path.stat().st_size
        self.workbook_values = load_workbook(self.file_path, read_only=True, data_only=True)
        self.workbook_formulas = load_workbook(self.file_path, read_only=True, data_only=False)
        try:
            warnings = []
            errors = []
            rows = []
            ignored_rows = 0
            rows_found = 0
            sheets_ignored_by_cutoff = 0
            eras_counter = Counter()
            sheets_report = {}
            validation = {
                "era_c": {
                    "biqueira": {"compared": 0, "correct": 0, "divergent": 0, "not_comparable": 0},
                    "approached_plus_reconductor": {"compared": 0, "correct": 0, "divergent": 0, "not_comparable": 0},
                }
            }

            reference_context = self._detect_reference_context()
            if reference_context.get("error"):
                errors.append(reference_context["error"])

            recognized = self._recognized_sheets()
            for sheet_name in recognized["daily"]:
                result = self._parse_daily_sheet(sheet_name, reference_context, validation)
                rows.extend(result["rows"])
                warnings.extend(result["warnings"])
                errors.extend(result["errors"])
                ignored_rows += result["rows_ignored"]
                rows_found += result["rows_found"]
                eras_counter[HistoricalTaxonomyEra.ERA_C] += len(result["rows"])
                sheets_report[sheet_name] = result["sheet_report"]
                if result["sheet_report"].get("ignored_by_cutoff"):
                    sheets_ignored_by_cutoff += 1

            for sheet_name in recognized["accumulated"]:
                result = self._parse_accumulated_sheet(sheet_name, reference_context)
                rows.extend(result["rows"])
                warnings.extend(result["warnings"])
                errors.extend(result["errors"])
                ignored_rows += result["rows_ignored"]
                rows_found += result["rows_found"]
                eras_counter[HistoricalTaxonomyEra.ERA_B] += len(result["rows"])
                sheets_report[sheet_name] = result["sheet_report"]

            for sheet_name in recognized["legacy"]:
                result = self._parse_legacy_sheet(sheet_name)
                rows.extend(result["rows"])
                warnings.extend(result["warnings"])
                errors.extend(result["errors"])
                ignored_rows += result["rows_ignored"]
                rows_found += result["rows_found"]
                eras_counter[HistoricalTaxonomyEra.ERA_A] += len(result["rows"])
                sheets_report[sheet_name] = result["sheet_report"]

            mother_report = self._inspect_mother_sheet()
            if mother_report:
                sheets_report[mother_report["sheet"]] = mother_report

            duplicate_batch_count = InspectionHistoricalImportBatch.objects.filter(source_file_sha256=file_sha256).count()
            if duplicate_batch_count:
                warnings.append(
                    f"Ja existe lote historico com o mesmo SHA-256 ({duplicate_batch_count} ocorrencia(s))."
                )

            date_values = sorted(row.reference_date for row in rows if row.reference_date is not None)
            teams = sorted({row.team for row in rows if row.team})
            reconciliation = self._build_reconciliation(rows, sheets_report)

            return {
                "file": {
                    "path": str(self.file_path),
                    "name": self.file_path.name,
                    "sha256": file_sha256,
                    "size": file_size,
                },
                "sheets": sheets_report,
                "summary": {
                    "rows_found": rows_found,
                    "rows_valid": len(rows),
                    "rows_imported": 0,
                    "rows_ignored": ignored_rows,
                    "errors": len(errors),
                    "warnings": len(warnings),
                    "sheets_ignored_by_cutoff": sheets_ignored_by_cutoff,
                    "historical_rows_by_era": dict(eras_counter),
                },
                "date_range": {
                    "start": date_values[0].isoformat() if date_values else None,
                    "end": date_values[-1].isoformat() if date_values else None,
                    "context_month": reference_context.get("month"),
                    "context_year": reference_context.get("year"),
                },
                "teams": teams,
                "eras": dict(eras_counter),
                "validation": validation,
                "reconciliation": reconciliation,
                "warnings": warnings[:25],
                "errors": errors[:25],
                "duplicate_file_detected": duplicate_batch_count > 0,
                "raw_rows": rows,
            }
        finally:
            if self.workbook_values is not None:
                self.workbook_values.close()
                self.workbook_values = None
            if self.workbook_formulas is not None:
                self.workbook_formulas.close()
                self.workbook_formulas = None

    def _recognized_sheets(self):
        daily = []
        accumulated = []
        legacy = []
        for sheet_name in self.workbook_values.sheetnames:
            normalized = _normalize_sheet_name(sheet_name)
            if DAILY_SHEET_PATTERN.match(sheet_name):
                daily.append(sheet_name)
            elif normalized.startswith("ACUMULADOS FISCALIZA"):
                accumulated.append(sheet_name)
            elif normalized == "PLAN2":
                legacy.append(sheet_name)
        return {"daily": daily, "accumulated": accumulated, "legacy": legacy}

    def _detect_reference_context(self):
        explicit_evidence = []
        generic_evidence = []
        for sheet_name in self.workbook_values.sheetnames:
            worksheet = self.workbook_values[sheet_name]
            for row in worksheet.iter_rows(min_row=1, max_row=min(40, worksheet.max_row), max_col=min(10, worksheet.max_column), values_only=True):
                for value in row:
                    if isinstance(value, datetime):
                        generic_evidence.append((value.year, value.month, f"{sheet_name}:datetime"))
                    elif isinstance(value, date):
                        generic_evidence.append((value.year, value.month, f"{sheet_name}:date"))
                    else:
                        text = _normalize_text(value).upper()
                        explicit_match = EXPLICIT_MONTH_YEAR_PATTERN.search(text)
                        if explicit_match:
                            month_name = explicit_match.group("month")
                            month_number = MONTH_NAMES.get(month_name)
                            if month_number:
                                explicit_evidence.append(
                                    (int(explicit_match.group("year")), month_number, f"{sheet_name}:explicit_text")
                                )
                        for month_name, month_number in MONTH_NAMES.items():
                            if month_name in text:
                                year_match = re.search(r"(20\d{2})", text)
                                if year_match:
                                    generic_evidence.append((int(year_match.group(1)), month_number, f"{sheet_name}:text"))

        if explicit_evidence:
            unique = {(year, month) for year, month, _source in explicit_evidence}
            if len(unique) == 1:
                year, month = next(iter(unique))
                return {"year": year, "month": month, "evidence": explicit_evidence[:10]}
            return {"error": f"Contexto mensal explicito ambiguo no workbook: {sorted(unique)}"}

        unique = {(year, month) for year, month, _source in generic_evidence}
        if len(unique) == 1:
            year, month = next(iter(unique))
            return {"year": year, "month": month, "evidence": generic_evidence[:10]}
        if not unique:
            return {"error": "Nao foi possivel identificar mes/ano de referencia do workbook."}
        return {"error": f"Contexto mensal ambiguo no workbook: {sorted(unique)}"}

    def _parse_daily_sheet(self, sheet_name, reference_context, validation):
        worksheet = self.workbook_values[sheet_name]
        warnings = []
        errors = []
        rows = []
        rows_found = 0
        rows_ignored = 0
        day_match = DAILY_SHEET_PATTERN.match(sheet_name)
        if not day_match:
            return {
                "rows": [],
                "warnings": [f"Aba diaria invalida: {sheet_name}"],
                "errors": [],
                "rows_ignored": 0,
                "rows_found": 0,
                "sheet_report": {"sheet": sheet_name, "source_type": "DAILY", "taxonomy_era": "ERA_C"},
            }
        if "year" not in reference_context or "month" not in reference_context:
            errors.append(f"Nao foi possivel derivar data segura para a aba {sheet_name}.")
            return {
                "rows": [],
                "warnings": warnings,
                "errors": errors,
                "rows_ignored": 0,
                "rows_found": 0,
                "sheet_report": {"sheet": sheet_name, "source_type": "DAILY", "taxonomy_era": "ERA_C", "errors": len(errors)},
            }

        day = int(day_match.group("day"))
        try:
            reference_date = date(reference_context["year"], reference_context["month"], day)
        except ValueError:
            errors.append(f"Data invalida derivada da aba {sheet_name}.")
            reference_date = None

        ignored_by_cutoff = False
        if reference_date and reference_date > HISTORICAL_CUTOFF_DATE:
            ignored_by_cutoff = True

        is_historical_daily = bool(reference_date and reference_date <= HISTORICAL_CUTOFF_DATE)

        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=4, max_row=min(1000, worksheet.max_row), max_col=23, values_only=True),
            start=4,
        ):
            source_team_label = _normalize_text(row[0])
            values = row[1:23]
            if not source_team_label:
                rows_ignored += 1
                continue
            if source_team_label.upper() in {"EQUIPE", "REGIÕES", "REGIOES"}:
                rows_ignored += 1
                continue
            if not source_team_label.upper().startswith("EQUIPE"):
                rows_ignored += 1
                continue
            if not any(value not in (None, "") for value in values):
                rows_ignored += 1
                continue

            rows_found += 1
            team, team_warning = self._normalize_team(source_team_label)
            if team_warning:
                warnings.append(f"{sheet_name}:{row_index} - {team_warning}")

            if not is_historical_daily:
                continue

            (
                approached_plus_reconductor,
                fined,
                towed,
                historical_cnh_retained,
                biqueira,
                historical_reconductors_licensed,
                refusal,
                four_ml,
                thirtythree_ml,
                thirtyfour_ml,
                arrests_means_evidence,
                taxi_approached,
                taxi_illegal,
                planned_actions,
                historical_deliberations,
                historical_event_trailers,
                rain,
                external_occurrence,
                public_security_occurrence,
                historical_operations,
                license_suspension,
                driving_canceled_license,
            ) = values

            self._validate_era_c_formula(
                validation["era_c"]["biqueira"],
                expected=_int_or_none(biqueira),
                calculated_components=[_int_or_none(four_ml), _int_or_none(thirtythree_ml), _int_or_none(thirtyfour_ml)],
            )
            self._validate_era_c_formula(
                validation["era_c"]["approached_plus_reconductor"],
                expected=_int_or_none(approached_plus_reconductor),
                calculated_components=[_int_or_none(four_ml), _int_or_none(refusal)],
            )

            rows.append(
                ParsedHistoricalRow(
                    source_type=HistoricalSourceType.DAILY,
                    taxonomy_era=HistoricalTaxonomyEra.ERA_C,
                    source_sheet=sheet_name,
                    source_row=row_index,
                    reference_date=reference_date,
                    reference_year=reference_date.year if reference_date else reference_context["year"],
                    reference_month=reference_date.month if reference_date else reference_context["month"],
                    team=team,
                    source_team_label=source_team_label,
                    source_workbook_label=self.workbook_label,
                    metrics={
                        "fined": _int_or_none(fined),
                        "towed": _int_or_none(towed),
                        "refusal": _int_or_none(refusal),
                        "four_ml": _int_or_none(four_ml),
                        "thirtythree_ml": _int_or_none(thirtythree_ml),
                        "thirtyfour_ml": _int_or_none(thirtyfour_ml),
                        "arrests_means_evidence": _int_or_none(arrests_means_evidence),
                        "taxi_approached": _int_or_none(taxi_approached),
                        "taxi_illegal": _int_or_none(taxi_illegal),
                        "planned_actions": _int_or_none(planned_actions),
                        "rain": _int_or_none(rain),
                        "external_occurrence": _int_or_none(external_occurrence),
                        "public_security_occurrence": _int_or_none(public_security_occurrence),
                        "license_suspension": _int_or_none(license_suspension),
                        "driving_canceled_license": _int_or_none(driving_canceled_license),
                        "historical_event_trailers": _int_or_none(historical_event_trailers),
                        "historical_reconductors_licensed": _int_or_none(historical_reconductors_licensed),
                        "historical_deliberations": _int_or_none(historical_deliberations),
                        "historical_operations": _int_or_none(historical_operations),
                        "historical_cnh_retained": _int_or_none(historical_cnh_retained),
                    },
                )
            )

        return {
            "rows": rows,
            "warnings": warnings,
            "errors": errors,
            "rows_ignored": rows_ignored,
            "rows_found": rows_found,
            "sheet_report": {
                "sheet": sheet_name,
                "source_type": HistoricalSourceType.DAILY,
                "taxonomy_era": HistoricalTaxonomyEra.ERA_C,
                "rows_found": rows_found,
                "rows_valid": len(rows),
                "rows_ignored": rows_ignored,
                "errors": len(errors),
                "warnings": len(warnings),
                "reference_date": reference_date.isoformat() if reference_date else None,
                "ignored_by_cutoff": ignored_by_cutoff,
            },
        }

    def _parse_accumulated_sheet(self, sheet_name, reference_context):
        worksheet = self.workbook_values[sheet_name]
        warnings = []
        errors = []
        rows = []
        rows_found = 0
        rows_ignored = 0
        current_month = None
        current_year = reference_context.get("year")

        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=min(700, worksheet.max_row), max_col=17, values_only=True),
            start=1,
        ):
            label = _normalize_text(row[0]).upper()
            if label in MONTH_NAMES:
                current_month = MONTH_NAMES[label]
                continue
            if not label.startswith("EQUIPE"):
                rows_ignored += 1
                continue
            values = row[1:17]
            if not any(value not in (None, "") for value in values):
                rows_ignored += 1
                continue
            rows_found += 1
            team, team_warning = self._normalize_team(_normalize_text(row[0]))
            if team_warning:
                warnings.append(f"{sheet_name}:{row_index} - {team_warning}")
            rows.append(
                ParsedHistoricalRow(
                    source_type=HistoricalSourceType.ACCUMULATED,
                    taxonomy_era=HistoricalTaxonomyEra.ERA_B,
                    source_sheet=sheet_name,
                    source_row=row_index,
                    reference_date=None,
                    reference_year=current_year,
                    reference_month=current_month,
                    team=team,
                    source_team_label=_normalize_text(row[0]),
                    source_workbook_label=self.workbook_label,
                    metrics={
                        "fined": _int_or_none(values[1]),
                        "towed": _int_or_none(values[2]),
                        "refusal": _int_or_none(values[6]),
                        "four_ml": _int_or_none(values[7]),
                        "thirtythree_ml": _int_or_none(values[8]),
                        "thirtyfour_ml": _int_or_none(values[9]),
                        "arrests_means_evidence": _int_or_none(values[10]),
                        "taxi_approached": _int_or_none(values[13]),
                        "taxi_illegal": _int_or_none(values[14]),
                        "rain": _int_or_none(values[15]),
                        "historical_reconductors_licensed": _int_or_none(values[5]),
                        "historical_cnh_retained": _int_or_none(values[3]),
                        "historical_alcohol_cases": _int_or_none(values[11]),
                        "historical_alcohol_percentage": _decimal_or_none(values[12]),
                    },
                )
            )

        return {
            "rows": rows,
            "warnings": warnings,
            "errors": errors,
            "rows_ignored": rows_ignored,
            "rows_found": rows_found,
            "sheet_report": {
                "sheet": sheet_name,
                "source_type": HistoricalSourceType.ACCUMULATED,
                "taxonomy_era": HistoricalTaxonomyEra.ERA_B,
                "rows_found": rows_found,
                "rows_valid": len(rows),
                "rows_ignored": rows_ignored,
                "errors": len(errors),
                "warnings": len(warnings),
            },
        }

    def _parse_legacy_sheet(self, sheet_name):
        worksheet = self.workbook_values[sheet_name]
        warnings = []
        errors = []
        rows = []
        rows_found = 0
        rows_ignored = 0

        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=3, max_row=min(500, worksheet.max_row), max_col=18, values_only=True),
            start=3,
        ):
            label = _normalize_text(row[0]).upper()
            if not label.startswith("EQUIPE"):
                rows_ignored += 1
                continue
            values = row[1:18]
            if not any(value not in (None, "") for value in values):
                rows_ignored += 1
                continue
            rows_found += 1
            team, team_warning = self._normalize_team(_normalize_text(row[0]))
            if team_warning:
                warnings.append(f"{sheet_name}:{row_index} - {team_warning}")
            rows.append(
                ParsedHistoricalRow(
                    source_type=HistoricalSourceType.LEGACY,
                    taxonomy_era=HistoricalTaxonomyEra.ERA_A,
                    source_sheet=sheet_name,
                    source_row=row_index,
                    reference_date=None,
                    reference_year=None,
                    reference_month=None,
                    team=team,
                    source_team_label=_normalize_text(row[0]),
                    source_workbook_label=self.workbook_label,
                    metrics={
                        "fined": _int_or_none(values[1]),
                        "towed": _int_or_none(values[2]),
                        "taxi_approached": _int_or_none(values[14]),
                        "taxi_illegal": _int_or_none(values[15]),
                        "rain": _int_or_none(values[16]),
                        "historical_cnh_retained": _int_or_none(values[3]),
                        "historical_passive_tests": _int_or_none(values[5]),
                        "historical_reconductors_licensed": _int_or_none(values[6]),
                        "refusal": _int_or_none(values[7]),
                        "negative_tests": _int_or_none(values[8]),
                        "administrative_art_165": _int_or_none(values[9]),
                        "criminal_art_306": _int_or_none(values[10]),
                        "criminal_art_306_other_evidence": _int_or_none(values[11]),
                        "historical_alcohol_cases": _int_or_none(values[12]),
                        "historical_alcohol_percentage": _decimal_or_none(values[13]),
                    },
                )
            )

        return {
            "rows": rows,
            "warnings": warnings,
            "errors": errors,
            "rows_ignored": rows_ignored,
            "rows_found": rows_found,
            "sheet_report": {
                "sheet": sheet_name,
                "source_type": HistoricalSourceType.LEGACY,
                "taxonomy_era": HistoricalTaxonomyEra.ERA_A,
                "rows_found": rows_found,
                "rows_valid": len(rows),
                "rows_ignored": rows_ignored,
                "errors": len(errors),
                "warnings": len(warnings),
            },
        }

    def _inspect_mother_sheet(self):
        for sheet_name in self.workbook_values.sheetnames:
            normalized = _normalize_sheet_name(sheet_name)
            if "(2)" in normalized and ("MÃE" in normalized or "MAE" in normalized):
                worksheet = self.workbook_values[sheet_name]
                non_empty_rows = 0
                for row in worksheet.iter_rows(min_row=1, max_row=min(40, worksheet.max_row), max_col=min(14, worksheet.max_column), values_only=True):
                    if any(value not in (None, "") for value in row):
                        non_empty_rows += 1
                return {
                    "sheet": sheet_name,
                    "source_type": "VALIDATION_ONLY",
                    "taxonomy_era": "VALIDATION_ONLY",
                    "rows_found": non_empty_rows,
                    "rows_valid": 0,
                    "rows_ignored": non_empty_rows,
                    "errors": 0,
                    "warnings": 0,
                    "used_for_import": False,
                }
        return None

    def _build_reconciliation(self, rows, sheets_report):
        daily_count = sum(1 for row in rows if row.source_type == HistoricalSourceType.DAILY)
        accumulated_count = sum(1 for row in rows if row.source_type == HistoricalSourceType.ACCUMULATED)
        mother_used = any(report.get("source_type") == "VALIDATION_ONLY" for report in sheets_report.values())
        status = "NOT_COMPARABLE"
        if daily_count and accumulated_count:
            status = "POTENTIAL_OVERLAP"
        return {
            "daily_vs_accumulated": {
                "status": status,
                "daily_rows": daily_count,
                "accumulated_rows": accumulated_count,
            },
            "mother_sheet": {
                "status": "VALIDATION_AVAILABLE" if mother_used else "MISSING",
            },
        }

    def _normalize_team(self, source_team_label):
        match = TEAM_PATTERN.match(source_team_label or "")
        if not match:
            return source_team_label, "Nao foi possivel normalizar a equipe de forma inequivoca."
        normalized = match.group(1).strip().replace(" -", "").strip()
        tokens = normalized.split()
        if not tokens:
            return source_team_label, "Nao foi possivel normalizar a equipe de forma inequivoca."

        first_token = tokens[0].strip()
        if re.fullmatch(r"[A-Z]+\d+", first_token) or re.fullmatch(r"\d+", first_token):
            if len(tokens) > 1 and re.fullmatch(r"[A-Z]", tokens[1]) and re.fullmatch(r"\d+", first_token):
                return f"{first_token} {tokens[1]}", None
            return first_token, None

        if re.fullmatch(r"[A-Z]+\d+[A-Z]", first_token):
            return first_token, None

        return normalized, "Nao foi possivel normalizar a equipe de forma inequivoca."

    def _validate_era_c_formula(self, stats_bucket, expected, calculated_components):
        if expected is None or any(component is None for component in calculated_components):
            stats_bucket["not_comparable"] += 1
            return
        stats_bucket["compared"] += 1
        calculated = sum(calculated_components)
        if calculated == expected:
            stats_bucket["correct"] += 1
        else:
            stats_bucket["divergent"] += 1


class InspectionHistoricalDryRunService:
    def run(self, file_path):
        parser = InspectionHistoricalWorkbookParser(file_path)
        return parser.parse()

    def render_report(self, file_path):
        report = self.run(file_path)
        report.pop("raw_rows", None)
        return json.dumps(report, ensure_ascii=False, indent=2, default=str)


class InspectionHistoricalImportService:
    def apply(self, file_path, source_type, taxonomy_era):
        from django.db import transaction
        from django.utils import timezone

        parser = InspectionHistoricalWorkbookParser(file_path)
        report = parser.parse()

        file_sha256 = report["file"]["sha256"]

        # Check if already imported
        existing_batch = InspectionHistoricalImportBatch.objects.filter(
            source_file_sha256=file_sha256,
            source_type=source_type,
            taxonomy_era=taxonomy_era,
            status=InspectionHistoricalImportBatch.Status.COMPLETED
        ).first()

        if existing_batch:
            return json.dumps({
                "error": f"Lote ja importado com sucesso para a fase {source_type}/{taxonomy_era} (Lote ID: {existing_batch.id})."
            })

        rows_to_import = [
            row for row in report.pop("raw_rows", [])
            if row.source_type == source_type and row.taxonomy_era == taxonomy_era
        ]

        batch = InspectionHistoricalImportBatch.objects.create(
            source_file_name=report["file"]["name"],
            source_file_sha256=file_sha256,
            source_file_size=report["file"]["size"],
            source_type=source_type,
            taxonomy_era=taxonomy_era,
            status=InspectionHistoricalImportBatch.Status.PENDING,
            started_at=timezone.now(),
            rows_found=len(rows_to_import),
            rows_valid=len(rows_to_import),
            errors_count=len(report["errors"]),
            warnings_count=len(report["warnings"]),
            report_json=report,
        )

        try:
            with transaction.atomic():
                stats = []
                for row in rows_to_import:
                    stats.append(
                        InspectionHistoricalStatistic(
                            reference_date=row.reference_date,
                            reference_year=row.reference_year,
                            reference_month=row.reference_month,
                            team=row.team,
                            source_team_label=row.source_team_label,
                            source_type=row.source_type,
                            source_sheet=row.source_sheet,
                            source_row=row.source_row,
                            taxonomy_era=row.taxonomy_era,
                            import_batch=batch,
                            source_workbook_label=row.source_workbook_label,
                            **row.metrics
                        )
                    )
                InspectionHistoricalStatistic.objects.bulk_create(stats)

                batch.rows_imported = len(stats)
                batch.status = InspectionHistoricalImportBatch.Status.COMPLETED
                batch.finished_at = timezone.now()
                batch.save()

                report["import_result"] = {
                    "batch_id": batch.id,
                    "rows_imported": len(stats),
                    "status": "COMPLETED",
                }

                return json.dumps(report, ensure_ascii=False, indent=2, default=str)
        except Exception as e:
            batch.status = InspectionHistoricalImportBatch.Status.FAILED
            batch.finished_at = timezone.now()
            batch.save()
            return json.dumps({
                "error": f"Erro durante a importacao: {str(e)}",
                "batch_id": batch.id,
            })
