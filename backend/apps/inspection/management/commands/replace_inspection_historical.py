import hashlib
import json
from collections import Counter
from datetime import date, datetime
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import Sum
from django.utils import timezone
from openpyxl import load_workbook

from apps.inspection.models import (
    HistoricalSourceType, HistoricalTaxonomyEra, InspectionHistoricalImportBatch,
    InspectionHistoricalStatistic, InspectionHistoricalTerritorialStatistic,
    InspectionMunicipality, InspectionReport, InspectionReportOperation, InspectionStatistic,
)
from apps.inspection.territorial import normalize_municipality_name


START, END = date(2023, 1, 1), date(2026, 8, 3)
PRESERVED_START, PRESERVED_END = date(2026, 8, 4), date(2026, 8, 9)
EXPECTED_ROWS = 11175
NUMERIC = {
    "Abordados": "historical_approached", "Multados": "fined", "Rebocados": "towed",
    "CNH recolhidas": "cnh_collected", "Testes com biqueira": "passive_tests_performed",
    "Recondutores": "reconductor", "Recusas": "refusal", "De 0,0 a 0,10": "four_ml",
    "De 0,11 a 0,29": "thirtythree_ml", "Mais de 0,30": "thirtyfour_ml",
    "Presos por outros motivos": "arrests_means_evidence", "Total de ações": "operations_count",
}
CONTROL = {2023: (3142, 3946, 299600, 109671, 6833, 1608, 264599, 43239, 33341, 34094),
           2024: (3094, 3463, 277754, 103627, 10, 202, 249288, 43047, 29478, 30552),
           2025: (3116, 3346, 354084, 117232, 353, 266, 326226, 43934, 29241, 30657),
           2026: (1823, 1855, 192111, 59054, 9, 86, 178644, 22033, 14205, 14953)}


def integer(value):
    if value in (None, ""): return 0
    value = float(value)
    if value != int(value): raise ValueError(f"valor inteiro inválido: {value}")
    return int(value)


class Command(BaseCommand):
    help = "Substitui com segurança o histórico DAILY/ERA_C de 2023-01-01 a 2026-08-03."
    expected_rows = EXPECTED_ROWS
    control_totals = CONTROL

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True)
        parser.add_argument("--dry-run", action="store_true")
        parser.add_argument("--apply", action="store_true")
        parser.add_argument("--confirm")

    def handle(self, *args, **opts):
        if opts["dry_run"] == opts["apply"]:
            raise CommandError("Informe exatamente um de --dry-run ou --apply.")
        if opts["apply"] and opts["confirm"] != "REPLACE-2023-2026":
            raise CommandError("--apply exige --confirm REPLACE-2023-2026.")
        rows, sha = self._read(opts["file"])
        report = self._report(rows, sha)
        if opts["dry_run"]:
            self.stdout.write(json.dumps(report, ensure_ascii=False, indent=2, default=str))
            return
        self._apply(rows, sha, Path(opts["file"]), report)
        self.stdout.write(self.style.SUCCESS("Substituição concluída com reconciliação integral."))

    def _read(self, path):
        file_path = Path(path)
        if not file_path.is_file(): raise CommandError(f"Arquivo não encontrado: {file_path}")
        sha = hashlib.sha256(file_path.read_bytes()).hexdigest()
        wb = load_workbook(file_path, read_only=True, data_only=True)
        if "Base Importacao" not in wb.sheetnames: raise CommandError("Aba 'Base Importacao' não encontrada.")
        sheet = wb["Base Importacao"]
        headers = [str(v).strip() if v is not None else "" for v in next(sheet.iter_rows(values_only=True))]
        required = {"Ano", "Data", "Equipe", "Município", "Apto para importação", "Pendências", "Alcoolemia calculada", *NUMERIC}
        missing = required - set(headers)
        if missing: raise CommandError(f"Colunas obrigatórias ausentes: {', '.join(sorted(missing))}")
        rows, errors, seen = [], [], set()
        for line, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            raw = dict(zip(headers, values)); raw["_line"] = line
            try:
                raw["Data"] = raw["Data"].date() if isinstance(raw["Data"], datetime) else raw["Data"]
                if not isinstance(raw["Data"], date) or not START <= raw["Data"] <= END: raise ValueError("Data fora do intervalo")
                if not str(raw["Equipe"] or "").strip(): raise ValueError("Equipe vazia")
                if not str(raw["Município"] or "").strip(): raise ValueError("Município vazio")
                if str(raw["Apto para importação"]).strip().lower() not in {"true", "sim", "1", "yes"}: raise ValueError("não apto")
                if str(raw["Pendências"] or "").strip(): raise ValueError("Pendências preenchidas")
                if integer(raw["Ano"]) != raw["Data"].year: raise ValueError("Ano diverge da Data")
                for column in NUMERIC: raw[column] = integer(raw[column])
                raw["Alcoolemia calculada"] = integer(raw["Alcoolemia calculada"])
                if any(raw[column] < 0 for column in NUMERIC) or raw["Alcoolemia calculada"] < 0: raise ValueError("indicador negativo")
                alcohol = raw["Recusas"] + raw["De 0,11 a 0,29"] + raw["Mais de 0,30"] + raw["Presos por outros motivos"]
                if raw["Alcoolemia calculada"] != alcohol: raise ValueError("Alcoolemia calculada diverge")
                key = (raw["Data"], str(raw["Equipe"]).strip().upper())
                if key in seen: raise ValueError("Data + Equipe duplicada")
                seen.add(key); rows.append(raw)
            except (TypeError, ValueError) as exc: errors.append(f"linha {line}: {exc}")
        wb.close()
        if errors or len(rows) != self.expected_rows: raise CommandError("; ".join(errors[:20] + ([] if len(rows) == self.expected_rows else [f"esperados {self.expected_rows}, encontrados {len(rows)}"])))
        totals = self._totals(rows)
        if any(totals[year] != self.control_totals[year] for year in self.control_totals): raise CommandError("Totais de controle da planilha divergem.")
        cities = {normalize_municipality_name(row["Município"]) for row in rows}
        known = set(InspectionMunicipality.objects.filter(is_active=True, region__is_active=True).values_list("normalized_name", flat=True))
        unknown = sorted(cities - known)
        if unknown: raise CommandError(f"Municípios não resolvidos: {', '.join(unknown)}")
        return rows, sha

    def _totals(self, rows):
        fields = ["Total de ações", "Abordados", "Multados", "Rebocados", "CNH recolhidas", "Testes com biqueira", "Recondutores", "Recusas", "Alcoolemia calculada"]
        return {year: (sum(r["Data"].year == year for r in rows), *(sum(r[f] for r in rows if r["Data"].year == year) for f in fields)) for year in self.control_totals}

    def _queryset(self):
        return InspectionHistoricalStatistic.objects.filter(reference_date__range=(START, END), source_type=HistoricalSourceType.DAILY, taxonomy_era=HistoricalTaxonomyEra.ERA_C, is_validation_only=False)

    def _report(self, rows, sha):
        current = self._queryset()
        territorial = InspectionHistoricalTerritorialStatistic.objects.filter(reference_date__range=(START, END))
        return {"sha256": sha, "valid_rows": len(rows), "totals_by_year": self._totals(rows), "would_delete": current.count(), "territorial_would_delete": territorial.count(), "preserved_2026_08_04_to_09": InspectionHistoricalStatistic.objects.filter(reference_date__range=(PRESERVED_START, PRESERVED_END)).count(), "projected_final_rows": InspectionHistoricalStatistic.objects.exclude(reference_date__range=(START, END)).count() + len(rows), "errors": [], "warnings": []}

    def _apply(self, rows, sha, file_path, report):
        filename = file_path.name
        operational_before = (InspectionReport.objects.count(), InspectionReportOperation.objects.count(), InspectionStatistic.objects.count())
        preserved_before = list(InspectionHistoricalStatistic.objects.filter(reference_date__range=(PRESERVED_START, PRESERVED_END)).values_list("pk", flat=True))
        with transaction.atomic():
            batch = InspectionHistoricalImportBatch.objects.create(source_file_name=filename, source_file_sha256=sha, source_type=HistoricalSourceType.DAILY, taxonomy_era=HistoricalTaxonomyEra.ERA_C, source_file_size=file_path.stat().st_size, status="PENDING", started_at=timezone.now(), rows_found=len(rows), rows_valid=len(rows), report_json=report)
            self._queryset().delete(); InspectionHistoricalTerritorialStatistic.objects.filter(reference_date__range=(START, END)).delete()
            municipalities = {m.normalized_name: m for m in InspectionMunicipality.objects.select_related("region").filter(is_active=True, region__is_active=True)}
            main, territorial = [], []
            for r in rows:
                metrics = {field: r[column] for column, field in NUMERIC.items()}
                team = str(r["Equipe"]).strip().upper(); municipality = municipalities[normalize_municipality_name(r["Município"])]
                main.append(InspectionHistoricalStatistic(reference_date=r["Data"], reference_year=r["Data"].year, reference_month=r["Data"].month, team=team, source_team_label=str(r["Equipe"]).strip(), source_type="DAILY", taxonomy_era="ERA_C", source_sheet=str(r.get("Aba de origem") or "Base Importacao"), source_row=integer(r.get("Linha de origem") or r["_line"]), source_workbook_label=filename, import_batch=batch, notes=f"Base tratada. Ajustes: {r.get('Ajustes realizados') or ''}", historical_operations=r["Total de ações"], **metrics))
                territorial.append(InspectionHistoricalTerritorialStatistic(reference_date=r["Data"], team=str(r["Equipe"]).strip(), source_city=str(r["Município"]).strip(), normalized_city=municipality.normalized_name, municipality=municipality, region=municipality.region, reports_count=1, approach=r["Abordados"], operations_count=r["Total de ações"], reconductor=r["Recondutores"], refusal=r["Recusas"], fined=r["Multados"], towed=r["Rebocados"], cnh_collected=r["CNH recolhidas"], four_ml=r["De 0,0 a 0,10"], thirtythree_ml=r["De 0,11 a 0,29"], thirtyfour_ml=r["Mais de 0,30"], passive_tests_performed=r["Testes com biqueira"], arrests_means_evidence=r["Presos por outros motivos"]))
            InspectionHistoricalStatistic.objects.bulk_create(main, batch_size=500); InspectionHistoricalTerritorialStatistic.objects.bulk_create(territorial, batch_size=500)
            if self._queryset().count() != len(rows) or InspectionHistoricalTerritorialStatistic.objects.filter(reference_date__range=(START, END)).count() != len(rows): raise CommandError("Reconciliação de quantidade falhou")
            if list(InspectionHistoricalStatistic.objects.filter(reference_date__range=(PRESERVED_START, PRESERVED_END)).values_list("pk", flat=True)) != preserved_before or operational_before != (InspectionReport.objects.count(), InspectionReportOperation.objects.count(), InspectionStatistic.objects.count()): raise CommandError("Dados preservados ou operacionais foram alterados")
            batch.status="COMPLETED"; batch.rows_imported=len(rows); batch.finished_at=timezone.now(); batch.save(update_fields=["status", "rows_imported", "finished_at"])
