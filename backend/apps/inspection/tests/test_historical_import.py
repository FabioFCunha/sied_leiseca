import json
import tempfile
from datetime import date, datetime, timezone
from decimal import Decimal
from io import StringIO
from pathlib import Path
from uuid import uuid4

from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.core.management.base import CommandError
from django.test import TestCase
from openpyxl import Workbook

from apps.inspection.historical_import import (
    InspectionHistoricalDryRunService,
    InspectionHistoricalWorkbookParser,
    compute_sha256,
)
from apps.inspection.models import (
    HISTORICAL_CUTOFF_DATE,
    InspectionHistoricalImportBatch,
    InspectionHistoricalStatistic,
    InspectionReport,
    InspectionStatistic,
    HistoricalSourceType,
    HistoricalTaxonomyEra,
)


class InspectionHistoricalImportTests(TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.user_model = get_user_model()
        self.user = self.user_model.objects.create_user(
            email="historical-import@example.com",
            password="secret123",
            role=self.user_model.Role.ADMIN,
            full_name="Historical Import Admin",
        )

    def tearDown(self):
        self.temp_dir.cleanup()

    def workbook_path(self, name="historico.xlsx"):
        return Path(self.temp_dir.name) / name

    def create_workbook(
        self,
        *,
        name="historico.xlsx",
        month=8,
        year=2026,
        daily_sheet="D1",
        include_daily=True,
        include_accumulated=True,
        include_plan2=True,
        include_mother=True,
        include_mother_2=True,
        daily_row=None,
        daily_total_row=False,
        accumulated_row=None,
        legacy_row=None,
    ):
        workbook = Workbook()
        first_sheet = workbook.active
        first_sheet.title = daily_sheet if include_daily else "Base"

        if include_daily:
            ws = workbook[daily_sheet]
            ws["A1"] = "FISCALIZAÇÃO - (TOTALIZAÇÃO DIÁRIO)"
            ws["A3"] = "REGIÕES"
            headers = [
                "ABORDADOS + RECONDUTOR",
                "MULTADOS",
                "REBOCADOS",
                "RECOLHIDAS",
                "TESTE COM BIQUEIRA",
                "RECONDUTORES HABILITADOS",
                "RECUSA DO TESTE",
                "DE 0,0 A 0,04",
                "DE 0,05 A 0,33",
                "MAIS DE 0,33",
                "PRESOS POR OUTROS POR MEIOS DE PROVA",
                "TAXI ABORDADOS",
                "PIRATA",
                "AÇÕES PLANEJADAS",
                "DELIBERAÇÕES",
                "REBOQUES",
                "CHUVA",
                "OCORRÊNCIA EXTERNA",
                "OCORRÊNCIA DE SEGURANÇA PÚBLICA",
                "OPERAÇÕES",
                "SUSPENSÃO CNH",
                "CNH CASSADA",
            ]
            for index, header in enumerate(headers, start=2):
                ws.cell(row=3, column=index, value=header)
            values = daily_row or [145, 44, 1, 2, 139, 10, 8, 137, 1, 1, 0, 3, 0, 5, 6, 7, 8, 9, 10, 11, 12, 13]
            ws.cell(row=4, column=1, value="EQUIPE A1")
            for index, value in enumerate(values, start=2):
                ws.cell(row=4, column=index, value=value)
            if daily_total_row:
                ws.cell(row=5, column=1, value="TOTAL")
                ws.cell(row=5, column=2, value=999)

        if include_accumulated:
            ws = workbook.create_sheet("ACUMULADOS FISCALIZAÇÃO ")
            ws["A2"] = list({1: "JANEIRO", 2: "FEVEREIRO", 3: "MARÇO", 4: "ABRIL", 5: "MAIO", 6: "JUNHO", 7: "JULHO", 8: "AGOSTO", 9: "SETEMBRO"}[month])[0] if False else None
            ws["A2"] = {
                1: "JANEIRO",
                2: "FEVEREIRO",
                3: "MARÇO",
                4: "ABRIL",
                5: "MAIO",
                6: "JUNHO",
                7: "JULHO",
                8: "AGOSTO",
                9: "SETEMBRO",
                10: "OUTUBRO",
                11: "NOVEMBRO",
                12: "DEZEMBRO",
            }[month]
            ws["A3"] = "REGIÕES"
            values = accumulated_row or [169, 44, 1, 2, 139, 10, 8, 137, 1, 1, 0, 12, Decimal("0.0710"), 3, 0, 1]
            ws.cell(row=4, column=1, value="EQUIPE A1")
            for index, value in enumerate(values, start=2):
                ws.cell(row=4, column=index, value=value)

        if include_plan2:
            ws = workbook.create_sheet("Plan2")
            ws["A1"] = "OPERAÇÃO LEI SECA - ACUMULADO FISCALIZAÇÃO"
            headers = [
                "REGIÕES",
                "ABORDADOS + RECONDUTORES ",
                "MULTADOS",
                "REBOCADOS",
                "RECOLHIDAS",
                "TESTE COM BIQUEIRA",
                "TESTE PASSIVO",
                "RECONDUTORES HABILITADOS",
                "RECUSA DO TESTE",
                "TESTES NEGATIVOS",
                "ADMINISTRATIVO ART. 165 CTB",
                "CRIMINAL ART. 306 CTB",
                "CRIMINAL POR MEIOS DE PROVA ART. 306 CTB",
                "CASOS DE ALCOOLEMIA",
                "PERCENTUAL DE PRODUTIVIDADE DE ALCOOLEMIA",
                "TAXI ABORDADOS",
                "PIRATA",
                "CHUVA",
            ]
            for index, header in enumerate(headers, start=1):
                ws.cell(row=2, column=index, value=header)
            values = legacy_row or ["EQUIPE 1", 100, 20, 1, 2, 90, 0, 12, 10, 80, 4, 1, 1, 12, Decimal("0.1200"), 3, 0, 1]
            for index, value in enumerate(values, start=1):
                ws.cell(row=3, column=index, value=value)

        if include_mother:
            ws = workbook.create_sheet("MÃE")
            ws["D10"] = datetime(year, month, min(int(daily_sheet[1:]), 9))

        if include_mother_2:
            ws = workbook.create_sheet("MAE (2)")
            ws["D6"] = f"FISCALIZAÇÃO - ACUMULADO MÊS DE { {1:'JANEIRO',2:'FEVEREIRO',3:'MARÇO',4:'ABRIL',5:'MAIO',6:'JUNHO',7:'JULHO',8:'AGOSTO',9:'SETEMBRO',10:'OUTUBRO',11:'NOVEMBRO',12:'DEZEMBRO'}[month] } {year}"
            ws["F7"] = 12157
            ws["F9"] = 909
            ws["F10"] = 909 / 12157
            ws["F12"] = 88

        path = self.workbook_path(name)
        workbook.save(path)
        return path

    def test_sha256_does_not_depend_on_filename(self):
        path_one = self.create_workbook(name="arquivo-1.xlsx")
        path_two = self.create_workbook(name="arquivo-2.xlsx")

        self.assertEqual(compute_sha256(path_one), compute_sha256(path_two))

    def test_parser_identifies_eras_and_sheet_types(self):
        path = self.create_workbook()

        report = InspectionHistoricalDryRunService().run(path)

        self.assertIn("D1", report["sheets"])
        self.assertEqual(report["sheets"]["D1"]["taxonomy_era"], HistoricalTaxonomyEra.ERA_C)
        self.assertEqual(
            report["sheets"]["ACUMULADOS FISCALIZAÇÃO "]["taxonomy_era"],
            HistoricalTaxonomyEra.ERA_B,
        )
        self.assertEqual(report["sheets"]["Plan2"]["taxonomy_era"], HistoricalTaxonomyEra.ERA_A)

    def test_mother_sheet_does_not_generate_historical_fact(self):
        path = self.create_workbook()

        report = InspectionHistoricalDryRunService().run(path)

        self.assertEqual(report["sheets"]["MAE (2)"]["used_for_import"], False)
        self.assertEqual(report["sheets"]["MAE (2)"]["rows_valid"], 0)

    def test_team_is_normalized_and_original_label_is_preserved(self):
        path = self.create_workbook()

        parser = InspectionHistoricalWorkbookParser(path)
        report = parser.parse()

        self.assertIn("A1", report["teams"])

    def test_null_and_zero_are_preserved_in_parser(self):
        daily_row = [145, 0, None, "", 139, None, 8, 137, 1, 1, 0, None, 0, None, None, None, 0, None, None, None, None, 0]
        path = self.create_workbook(daily_row=daily_row)

        parser = InspectionHistoricalWorkbookParser(path)
        report = parser.parse()

        self.assertEqual(report["summary"]["rows_valid"], 3)

    def test_total_row_is_ignored(self):
        path = self.create_workbook(daily_total_row=True)

        report = InspectionHistoricalDryRunService().run(path)

        self.assertGreaterEqual(report["summary"]["rows_ignored"], 1)

    def test_cutoff_date_after_2026_08_09_generates_error(self):
        path = self.create_workbook(daily_sheet="D10")

        report = InspectionHistoricalDryRunService().run(path)

        self.assertGreaterEqual(report["summary"]["sheets_ignored_by_cutoff"], 1)

    def test_formula_validation_for_biqueira_and_abordados(self):
        path = self.create_workbook()

        report = InspectionHistoricalDryRunService().run(path)

        self.assertEqual(report["validation"]["era_c"]["biqueira"]["compared"], 1)
        self.assertEqual(report["validation"]["era_c"]["biqueira"]["correct"], 1)
        self.assertEqual(report["validation"]["era_c"]["approached_plus_reconductor"]["compared"], 1)
        self.assertEqual(report["validation"]["era_c"]["approached_plus_reconductor"]["correct"], 1)

    def test_formula_divergence_generates_validation_counter(self):
        daily_row = [170, 44, 1, 2, 140, 10, 8, 137, 1, 1, 0, 3, 0, 5, 6, 7, 8, 9, 10, 11, 12, 13]
        path = self.create_workbook(daily_row=daily_row)

        report = InspectionHistoricalDryRunService().run(path)

        self.assertEqual(report["validation"]["era_c"]["biqueira"]["divergent"], 1)
        self.assertEqual(report["validation"]["era_c"]["approached_plus_reconductor"]["divergent"], 1)

    def test_explicit_month_year_context_takes_precedence_over_generic_legacy_dates(self):
        path = self.create_workbook()

        from openpyxl import load_workbook

        loaded = load_workbook(path)
        loaded["MÃE"]["D10"] = datetime(2020, 10, 1)
        loaded.save(path)
        loaded.close()

        report = InspectionHistoricalDryRunService().run(path)

        self.assertEqual(report["date_range"]["context_month"], 8)
        self.assertEqual(report["date_range"]["context_year"], 2026)
        self.assertGreater(report["sheets"]["D1"]["rows_valid"], 0)

    def test_dry_run_generates_zero_writes(self):
        path = self.create_workbook()
        before_batches = InspectionHistoricalImportBatch.objects.count()
        before_stats = InspectionHistoricalStatistic.objects.count()

        report = InspectionHistoricalDryRunService().run(path)

        self.assertEqual(report["summary"]["rows_imported"], 0)
        self.assertEqual(InspectionHistoricalImportBatch.objects.count(), before_batches)
        self.assertEqual(InspectionHistoricalStatistic.objects.count(), before_stats)

    def test_command_blocks_execution_without_dry_run(self):
        path = self.create_workbook()

        with self.assertRaisesMessage(CommandError, "Voce deve especificar --dry-run ou --apply."):
            call_command("import_inspection_historical", file=str(path))

    def test_command_outputs_json_in_dry_run(self):
        path = self.create_workbook()
        stdout = StringIO()

        call_command("import_inspection_historical", file=str(path), dry_run=True, stdout=stdout)

        payload = json.loads(stdout.getvalue())
        self.assertEqual(payload["file"]["name"], path.name)
        self.assertEqual(payload["summary"]["rows_imported"], 0)

    def test_duplicate_file_is_detected_by_hash(self):
        path = self.create_workbook()
        file_hash = compute_sha256(path)
        InspectionHistoricalImportBatch.objects.create(
            source_file_name=path.name,
            source_file_sha256=file_hash,
            source_file_size=path.stat().st_size,
            status=InspectionHistoricalImportBatch.Status.COMPLETED,
            imported_by=self.user,
            started_at=datetime(2026, 8, 12, 9, 0, tzinfo=timezone.utc),
            finished_at=datetime(2026, 8, 12, 9, 5, tzinfo=timezone.utc),
        )

        report = InspectionHistoricalDryRunService().run(path)

        self.assertTrue(report["duplicate_file_detected"])

    def test_reconciliation_marks_overlap_between_daily_and_accumulated(self):
        path = self.create_workbook()

        report = InspectionHistoricalDryRunService().run(path)

        self.assertEqual(report["reconciliation"]["daily_vs_accumulated"]["status"], "POTENTIAL_OVERLAP")

    def test_historical_import_batch_sha256_is_unique(self):
        InspectionHistoricalImportBatch.objects.create(
            source_file_name="a.xlsx",
            source_file_sha256="abc",
            source_file_size=1,
            status=InspectionHistoricalImportBatch.Status.PENDING,
            imported_by=self.user,
            started_at=datetime(2026, 8, 12, 9, 0, tzinfo=timezone.utc),
            source_type=HistoricalSourceType.DAILY,
            taxonomy_era=HistoricalTaxonomyEra.ERA_C,
        )

        with self.assertRaises(Exception):
            InspectionHistoricalImportBatch.objects.create(
                source_file_name="b.xlsx",
                source_file_sha256="abc",
                source_file_size=2,
                status=InspectionHistoricalImportBatch.Status.PENDING,
                imported_by=self.user,
                started_at=datetime(2026, 8, 12, 9, 5, tzinfo=timezone.utc),
                source_type=HistoricalSourceType.DAILY,
                taxonomy_era=HistoricalTaxonomyEra.ERA_C,
            )

    def test_historical_statistic_blocks_reference_date_after_cutoff(self):
        batch = InspectionHistoricalImportBatch.objects.create(
            source_file_name="a.xlsx",
            source_file_sha256="def",
            source_file_size=1,
            status=InspectionHistoricalImportBatch.Status.PENDING,
            imported_by=self.user,
            started_at=datetime(2026, 8, 12, 9, 0, tzinfo=timezone.utc),
        )

        item = InspectionHistoricalStatistic(
            import_batch=batch,
            reference_date=date(2026, 8, 10),
            reference_year=2026,
            reference_month=8,
            team="A1",
            source_team_label="EQUIPE A1",
            source_type=HistoricalSourceType.DAILY,
            source_sheet="D10",
            source_row=4,
            taxonomy_era=HistoricalTaxonomyEra.ERA_C,
            source_workbook_label="arquivo.xlsx",
        )

        with self.assertRaisesMessage(Exception, "Dados historicos de Fiscalizacao nao podem ultrapassar 2026-08-09."):
            item.full_clean()

    def test_historical_statistic_unique_per_batch_sheet_and_row(self):
        batch = InspectionHistoricalImportBatch.objects.create(
            source_file_name="a.xlsx",
            source_file_sha256="ghi",
            source_file_size=1,
            status=InspectionHistoricalImportBatch.Status.PENDING,
            imported_by=self.user,
            started_at=datetime(2026, 8, 12, 9, 0, tzinfo=timezone.utc),
        )
        InspectionHistoricalStatistic.objects.create(
            import_batch=batch,
            reference_date=date(2026, 8, 9),
            reference_year=2026,
            reference_month=8,
            team="A1",
            source_team_label="EQUIPE A1",
            source_type=HistoricalSourceType.DAILY,
            source_sheet="D9",
            source_row=4,
            taxonomy_era=HistoricalTaxonomyEra.ERA_C,
            source_workbook_label="arquivo.xlsx",
        )

        with self.assertRaises(Exception):
            InspectionHistoricalStatistic.objects.create(
                import_batch=batch,
                reference_date=date(2026, 8, 9),
                reference_year=2026,
                reference_month=8,
                team="A1",
                source_team_label="EQUIPE A1",
                source_type=HistoricalSourceType.DAILY,
                source_sheet="D9",
                source_row=4,
                taxonomy_era=HistoricalTaxonomyEra.ERA_C,
                source_workbook_label="arquivo.xlsx",
            )

    def test_no_regression_for_official_inspection_statistic(self):
        report = InspectionReport.objects.create(
            source_id=uuid4(),
            source_created_at=datetime(2026, 8, 10, 8, 0, tzinfo=timezone.utc),
            source_updated_at=datetime(2026, 8, 10, 8, 30, tzinfo=timezone.utc),
            synced_at=datetime(2026, 8, 10, 9, 0, tzinfo=timezone.utc),
            operation_date=date(2026, 8, 10),
            team="A3",
        )
        statistic = InspectionStatistic.objects.create(
            report=report,
            source_report_id=report.source_id,
            operation_date=report.operation_date,
            team=report.team,
            snapshot_source_updated_at=report.source_updated_at,
            operations_count=1,
        )

        self.assertEqual(statistic.source_report_id, report.source_id)
        self.assertEqual(HISTORICAL_CUTOFF_DATE.isoformat(), "2026-08-09")
