import json
import re
from pathlib import Path

from openpyxl import load_workbook


visual_dir = Path(
    r"C:\Users\FABIO CUNHA\.codex\visualizations\2026\07\24"
    r"\019f93fd-4a0d-7321-b222-621e212ce8fc"
)
source_path = visual_dir / "painel-historico-completo-desde-2011.html"
target_path = visual_dir / "painel-historico-todos-indicadores.html"
workbook_path = Path(r"E:\agenda_eventos_ols\estatisticas.xlsx")

html = source_path.read_text(encoding="utf-8")
worksheet = load_workbook(workbook_path, data_only=True, read_only=True).active

years = [worksheet.cell(2, column).value for column in range(4, 20)]
labels = {
    3: "1 - Público total",
    4: "1.1 - Público em palestras",
    5: "1.2 - Público em ações",
    6: "2 - Palestras realizadas",
    7: "2.1 - Escolas",
    8: "2.2 - Universidades",
    9: "2.3 - Empresas",
    10: "3 - Ações",
    11: "3.1 - Bares",
    12: "3.2 - Pedágio",
    13: "3.3 - Esportes",
    14: "3.4 - Praia",
    15: "3.5 - Eventos",
    16: "3.6 - Shopping/Centro Comercial",
    17: "3.7 - Ação Social",
    18: "3.8 - Outros",
    19: "4 - Materiais de divulgação",
    20: "2.4 - Certificados entregues",
    21: "Revistinha Soprinho",
}
july_additions = {
    3: 2802,
    4: 252,
    5: 2550,
    6: 4,
    7: 2,
    9: 2,
    10: 25,
    15: 1,
    18: 24,
    19: 489,
    21: 439,
}

indicators = []
for row, label in labels.items():
    values = [int(worksheet.cell(row, column).value or 0) for column in range(4, 20)]
    values[-1] += july_additions.get(row, 0)
    indicators.append({"label": label, "values": values})

indicators.extend(
    [
        {"label": "3.9 - Praças/Parques Públicos (novo processo)", "values": [0] * 16},
        {"label": "3.10 - Pontos turísticos (novo processo)", "values": [0] * 16},
    ]
)

table_section = """
  <section>
    <h3>Série histórica completa de todos os indicadores</h3>
    <p class="text-muted">Todos os indicadores da planilha estatisticas.xlsx, de 2011 a 2026, com julho/SIED incorporado à coluna de 2026 e sem comparação percentual.</p>
    <div class="table-responsive">
      <table class="table table-sm indicator-table">
        <thead id="ols-all-indicators-head"></thead>
        <tbody id="ols-all-indicators-body"></tbody>
      </table>
    </div>
  </section>"""

html = re.sub(
    r"\s*<section>\s*<h3>Série histórica completa dos indicadores</h3>.*?</section>",
    "\n" + table_section,
    html,
    count=1,
    flags=re.DOTALL,
)

html = re.sub(
    r"\s*const completeHistoryBody = root\.querySelector\('#ols-complete-history-body'\);"
    r".*?(?=\s*const body = root\.querySelector\('#ols-options-body'\);)",
    "\n",
    html,
    count=1,
    flags=re.DOTALL,
)

dataset_script = f"""
      const allIndicatorYears = {json.dumps(years, ensure_ascii=False)};
      const allIndicators = {json.dumps(indicators, ensure_ascii=False, separators=(',', ':'))};
      const allIndicatorsHead = root.querySelector('#ols-all-indicators-head');
      const allIndicatorsBody = root.querySelector('#ols-all-indicators-body');
      allIndicatorsHead.innerHTML = `<tr><th>Indicador</th>${{allIndicatorYears.map(year => `<th class="text-end">${{year}}</th>`).join('')}}<th class="text-end">Total</th></tr>`;
      allIndicators.forEach(indicator => {{
        const row = document.createElement('tr');
        const total = indicator.values.reduce((sum, value) => sum + Number(value || 0), 0);
        row.innerHTML = `<td class="text-nowrap">${{indicator.label}}</td>${{indicator.values.map(value => `<td class="text-end">${{format(value)}}</td>`).join('')}}<td class="text-end">${{format(total)}}</td>`;
        allIndicatorsBody.appendChild(row);
      }});
"""

html = html.replace(
    "      const body = root.querySelector('#ols-options-body');",
    dataset_script + "\n      const body = root.querySelector('#ols-options-body');",
)

target_path.write_text(html, encoding="utf-8")
